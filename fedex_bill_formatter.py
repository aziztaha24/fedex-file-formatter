import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

st.set_page_config(page_title="FedEx Bill Formatter", page_icon="📦", layout="centered")

st.title("📦 FedEx Bill Formatter")
st.markdown("Upload your raw FedEx bill and download the formatted version instantly.")

def parse_date(val):
    if pd.isna(val):
        return None
    try:
        s = str(int(val))
        return datetime.strptime(s, "%Y%m%d").date()
    except:
        return None

def format_fedex_bill(uploaded_file):
    df = pd.read_excel(uploaded_file, sheet_name='Source File Bill')

    charge_desc_cols = []
    for col in df.columns:
        if 'Tracking ID Charge Description' in col:
            amt_col = col.replace('Description', 'Amount')
            if amt_col in df.columns:
                charge_desc_cols.append((col, amt_col))

    output_rows = []

    for _, row in df.iterrows():
        invoice_date  = parse_date(row.get('Invoice Date'))
        shipment_date = parse_date(row.get('Shipment Date'))
        customer_ref  = row.get('Original Customer Reference')
        tracking_id   = row.get('Express or Ground Tracking ID')
        invoice_no    = row.get('Invoice Number')
        invoice_total = row.get('Original Amount Due')
        service_type  = row.get('Service Type') if pd.notna(row.get('Service Type')) else row.get('Ground Service')
        transport_amt = row.get('Transportation Charge Amount')

        # Extract just the number from Package ID (e.g. "PKG ID: 94201" -> 94201)
        pkg_id_raw = row.get('Original Ref#2')
        pkg_id = None
        if pd.notna(pkg_id_raw):
            try:
                pkg_id = int(''.join(filter(str.isdigit, str(pkg_id_raw))))
            except:
                pkg_id = pkg_id_raw

        # Convert tracking ID to integer
        try:
            tracking_id = int(float(tracking_id)) if pd.notna(tracking_id) else tracking_id
        except:
            pass

        # Collect all charges
        charges = []
        for desc_col, amt_col in charge_desc_cols:
            desc = row.get(desc_col)
            amt  = row.get(amt_col)
            if pd.notna(desc) and pd.notna(amt):
                charges.append((str(desc).strip(), float(amt)))

        # Subtract negative charges from transportation amount
        adj_transport = transport_amt
        if pd.notna(transport_amt):
            for desc, amt in charges:
                if amt < 0:
                    adj_transport = round(adj_transport + amt, 2)

        common = {
            'Shipment Date': shipment_date,
            'Customer Ref.': customer_ref if pd.notna(customer_ref) else None,
            'Tracking ID': tracking_id,
            'Package ID': pkg_id,
            'Invoice Date': invoice_date,
            'Invoice No.': int(invoice_no) if pd.notna(invoice_no) else invoice_no,
            'Invoice Total': invoice_total,
        }

        if pd.notna(transport_amt):
            main_service = service_type if pd.notna(service_type) else None
            output_rows.append({**common, 'Service Type': main_service, 'Amount': adj_transport})
            for desc, amt in charges:
                if amt > 0:
                    output_rows.append({**common, 'Service Type': desc, 'Amount': amt})
        else:
            for desc, amt in charges:
                if amt > 0:
                    output_rows.append({**common, 'Service Type': desc, 'Amount': amt})

    result_df = pd.DataFrame(output_rows, columns=[
        'Shipment Date', 'Customer Ref.', 'Tracking ID', 'Package ID',
        'Service Type', 'Invoice Date', 'Invoice No.', 'Invoice Total', 'Amount'
    ])

    # Write to in-memory Excel file
    output = io.BytesIO()
    result_df.to_excel(output, index=False, sheet_name='Formatted')
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(name='Arial', bold=True, color="FFFFFF", size=10)
    data_font    = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align   = Alignment(horizontal='left',   vertical='center')
    thin_border  = Border(bottom=Side(style='thin', color='CCCCCC'))

    col_widths = {
        'A': 16, 'B': 22, 'C': 18, 'D': 14,
        'E': 34, 'F': 14, 'G': 14, 'H': 14, 'I': 12
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 18
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 16
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.font   = data_font
            cell.border = thin_border
            col_letter  = get_column_letter(col_idx)

            if col_letter in ('A', 'F'):         # Shipment Date, Invoice Date
                if cell.value is not None:
                    cell.number_format = 'MM/DD/YYYY'
                cell.alignment = center_align
            elif col_letter in ('C', 'D', 'G'):  # Tracking ID, Package ID, Invoice No.
                cell.number_format = '0'
                cell.alignment = center_align
            elif col_letter in ('H', 'I'):        # Invoice Total, Amount
                cell.number_format = '#,##0.00'
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    ws.freeze_panes = 'A2'

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output, result_df


# --- UI ---
uploaded_file = st.file_uploader("Upload FedEx Bill (.xlsx)", type=["xlsx"])

if uploaded_file:
    with st.spinner("Processing..."):
        try:
            output_bytes, preview_df = format_fedex_bill(uploaded_file)

            st.success(f"✅ Done! {len(preview_df)} rows generated.")
            st.dataframe(preview_df, use_container_width=True)

            st.download_button(
                label="⬇️ Download Formatted Excel",
                data=output_bytes,
                file_name="FedEx_Bill_Formatted.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ Error processing file: {e}")
